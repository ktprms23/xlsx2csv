from pathlib import Path 
from openpyxl import load_workbook





def get_file_name(data_dir):
 
    data_dir_path = Path(data_dir)
    for data_path in data_dir_path.glob('*.xlsx'):
        yield str(data_path)#'AOA/20190116_1M_0.csv'

def test_excel():

    wb = load_workbook('test.xlsx')
    print(wb.sheetnames)

    for line in wb.active.rows:
        print(line[0].value)
        break

def convertXLSX2CSV(folderName = 'folder'):

    data_filename = get_file_name(folderName)

    
#    print(data_filename_a)
#    print(data_filename_b)
#    print(data_filename_c)
#    print(data_filename_d)
#    exit()
#t = time.time()
#print(t)


    for file_name in data_filename:
        print( 'parse file:' + file_name )
        outputFile = open( file_name + '.csv', "w", encoding='utf8')
        #outputFile = open( file_name, "r", encoding='utf8')
        wb = load_workbook(file_name)
        
#    outputFile = open( '20190116_1M_0.csv', "r", encoding='utf8')
#        lines = [line.rstrip('\n') for line in outputFile]
        lines = wb.active.rows
        for line in lines:
        
            line_csv = ""
            index = 0
            for s in line:
                if index != 0:
                    line_csv += ','
                else:
                    index = 1
 
                # if the value is empty, the value will be None, convert to ''
                if str(s.value) == 'None':
                    line_csv += ''
                else:
                    line_csv += str(s.value)
                
            outputFile.write(line_csv + '\n')

if __name__ == '__main__':
#    test_excel()
    convertXLSX2CSV("small_test/A")
    convertXLSX2CSV("small_test/B")
    convertXLSX2CSV("small_test/C")
    convertXLSX2CSV("small_test/D")

